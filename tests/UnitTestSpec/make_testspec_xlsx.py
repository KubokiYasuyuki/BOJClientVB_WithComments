#!/usr/bin/env python3
"""Generate BOJClientVB Unit Test Specification Excel file."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles -----------------------------------------------------------------
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="D9E1F2")
section_font = Font(name="Meiryo UI", size=11, bold=True, color="1F3864")
normal_font = Font(name="Meiryo UI", size=10)
title_font = Font(name="Meiryo UI", size=16, bold=True, color="1F3864")
subtitle_font = Font(name="Meiryo UI", size=11, bold=True)
wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_header_row(ws, row, headers, widths):
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_row(ws, row, values, alignments=None, border_on=True):
    for col_idx, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = normal_font
        if alignments and col_idx - 1 < len(alignments):
            cell.alignment = alignments[col_idx - 1]
        else:
            cell.alignment = wrap_left
        if border_on:
            cell.border = border


# =============================================================================
# Sheet 1: 表紙・概要
# =============================================================================
ws1 = wb.active
ws1.title = "表紙"
ws1.sheet_view.showGridLines = False

ws1["B2"] = "BOJClientVB 単体テスト仕様書"
ws1["B2"].font = title_font
ws1.merge_cells("B2:G2")

ws1["B4"] = "対象メソッド"
ws1["C4"] = "BOJClientVB.BOJClientVBCLS.AAVerifyID (COM 公開)"
ws1["B5"] = "対応ソース"
ws1["C5"] = "BOJClientVBCLS.vb (PR #1 マージ後)"
ws1["B6"] = "設計書"
ws1["C6"] = "ID-05 共通制御系処理機能仕様書 (2026.04.20 版)"
ws1["B7"] = "テスト方式"
ws1["C7"] = "結合カバレッジ (VBScript + スタブサーバ / 64bit 構成のみ)"
ws1["B8"] = "作成日"
ws1["C8"] = "2026/04/20"
ws1["B9"] = "版"
ws1["C9"] = "v1.0"

for r in range(4, 10):
    ws1.cell(row=r, column=2).font = subtitle_font
    ws1.cell(row=r, column=3).font = normal_font
    ws1.cell(row=r, column=2).alignment = wrap_left
    ws1.cell(row=r, column=3).alignment = wrap_left
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

ws1["B11"] = "1. テスト方針"
ws1["B11"].font = section_font
ws1["B11"].fill = section_fill
ws1.merge_cells("B11:G11")
ws1["B12"] = (
    "AAVerifyID は COM 公開されるため、実運用に近い形で VBScript から呼び出し、"
    "スタブサーバを相手にして szAccess / szPassword / online / szRet / szError の各出力を検証する。"
    "関数戻り値は設計書準拠で常に 0 とし、処理結果は szRet で判定する。"
)
ws1["B12"].font = normal_font
ws1["B12"].alignment = wrap_left
ws1.merge_cells("B12:G14")
ws1.row_dimensions[12].height = 50

ws1["B16"] = "2. szRet 値定義"
ws1["B16"].font = section_font
ws1["B16"].fill = section_fill
ws1.merge_cells("B16:G16")

set_header_row(ws1, 17, ["szRet", "意味"], [10, 60])
for i, (v, m) in enumerate([
    ("0", "正常終了"),
    ("1", "初期化失敗 (レジストリ読込失敗)"),
    ("2", "コネクション失敗 (Host01 / Host02 共に接続不可)"),
    ("3", "その他のエラー (サーバ NG 応答 / 応答解析失敗 等)"),
    ("4", "受信タイムアウト"),
]):
    write_row(ws1, 18 + i, [v, m], [wrap_center, wrap_left])

ws1["B24"] = "3. 本書のシート構成"
ws1["B24"].font = section_font
ws1["B24"].fill = section_fill
ws1.merge_cells("B24:G24")

set_header_row(ws1, 25, ["シート名", "内容"], [20, 50])
sheets_info = [
    ("表紙", "本シート。対象・方針・全体概要"),
    ("環境設定", "レジストリ・services ファイル・VBScript 呼出し用コード"),
    ("スタブサーバ", "PowerShell 製スタブサーバの仕様・モード・起動手順"),
    ("テストケース一覧", "AAVerifyID の全テストケース (15 件)"),
    ("実施記録", "結果記録 (実施日・実施者・結果・備考)"),
]
for i, (name, desc) in enumerate(sheets_info):
    write_row(ws1, 26 + i, [name, desc], [wrap_center, wrap_left])


# =============================================================================
# Sheet 2: 環境設定
# =============================================================================
ws2 = wb.create_sheet("環境設定")
ws2.sheet_view.showGridLines = False

ws2["B2"] = "環境設定"
ws2["B2"].font = title_font
ws2.merge_cells("B2:F2")

ws2["B4"] = "1. 基本環境"
ws2["B4"].font = section_font
ws2["B4"].fill = section_fill
ws2.merge_cells("B4:F4")

set_header_row(ws2, 5, ["区分", "内容"], [25, 60])
env_items = [
    ("クライアント OS", "Windows 10 / 11 (64bit)"),
    ("ビルド構成", "Visual Studio 2015 以降、Release / x64 (64bit) のみ。AnyCPU / x86 は対象外"),
    (".NET Framework", "4.5 以上"),
    ("COM 登録", "64bit 版 regasm で登録 : %WINDIR%\\Microsoft.NET\\Framework64\\v4.0.30319\\RegAsm.exe BOJClientVB.dll /codebase /tlb (/tlb は型ライブラリ登録用。遅延バインディングのみなら省略可だが付与しても害なし)"),
    ("呼び出し元", "64bit 版 cscript : C:\\Windows\\System32\\cscript.exe (デフォルト)。SysWOW64 側の 32bit cscript は使用しない"),
    ("呼び出し想定", "本アセンブリは Access 64bit 版からの呼び出し用。32bit アプリからの呼び出しはテスト対象外"),
    ("サーバ", "本番 xinetd は停止不要。スタブサーバをクライアントと同一 PC で起動"),
    ("管理者権限", "PowerShell / Visual Studio は「管理者として実行」で起動すること"),
]
for i, (k, v) in enumerate(env_items):
    write_row(ws2, 6 + i, [k, v])

ws2["B15"] = "2. バックアップ取得 (レジストリ / services 変更の前に必ず実施)"
ws2["B15"].font = section_font
ws2["B15"].fill = section_fill
ws2.merge_cells("B15:F15")

ws2["B16"] = "レジストリ : reg export \"HKLM\\SOFTWARE\\sys148\" sys148_backup.reg /reg:64"
ws2["B17"] = "services  : copy C:\\Windows\\System32\\drivers\\etc\\services services_backup.txt"
for r in [16, 17]:
    ws2.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws2["B19"] = "3. レジストリ設定 (既定値) ※ 64bit プロセスから参照するため WOW6432Node は使用しない"
ws2["B19"].font = section_font
ws2["B19"].fill = section_fill
ws2.merge_cells("B19:F19")

ws2["B20"] = "パス: HKEY_LOCAL_MACHINE\\SOFTWARE\\sys148\\Profiles"
ws2["B20"].font = normal_font
ws2.merge_cells("B20:F20")

set_header_row(ws2, 21, ["値名", "型", "値", "備考"], [18, 12, 18, 45])
reg_items = [
    ("Host01", "REG_SZ", "127.0.0.1", "スタブサーバを同一 PC で起動するため"),
    ("Host02", "REG_SZ", "127.0.0.1", "フェイルオーバ動作確認時は別値に変更"),
    ("ServiceName", "REG_SZ", "aauth", "services ファイル検索用"),
    ("SocketTimeOut", "REG_DWORD", "5", "受信タイムアウト検証のため短めに設定"),
]
for i, row in enumerate(reg_items):
    write_row(ws2, 22 + i, list(row), [wrap_left, wrap_center, wrap_center, wrap_left])

ws2.cell(row=26, column=2, value=(
    "※ regedit で確認する場合は 64bit 版 regedit (C:\\Windows\\regedit.exe) を使用する。"
    "SOFTWARE\\WOW6432Node 配下は 32bit 用リダイレクト先のため本テストでは参照しない。"
))
ws2.cell(row=26, column=2).font = normal_font
ws2.cell(row=26, column=2).alignment = wrap_left
ws2.merge_cells("B26:F26")

ws2["B28"] = "4. services ファイル設定"
ws2["B28"].font = section_font
ws2["B28"].fill = section_fill
ws2.merge_cells("B28:F28")

ws2["B29"] = "パス: C:\\Windows\\System32\\drivers\\etc\\services"
ws2["B29"].font = normal_font
ws2.merge_cells("B29:F29")

ws2["B30"] = "追記内容: aauth    50000/tcp"
ws2["B30"].font = normal_font
ws2.merge_cells("B30:F30")

ws2["B32"] = "5. VBScript 呼出しコード (test_aaverifyid.vbs)"
ws2["B32"].font = section_font
ws2["B32"].fill = section_fill
ws2.merge_cells("B32:F32")

vbs_code = """Set obj = CreateObject("BOJClientVB.BOJClientVBCLS")
Dim szAccess, szPassword, online, szRet, szError
Dim t0, t1
t0 = Timer
ret = obj.AAVerifyID("DUMMY", "DUMMY", szAccess, szPassword, online, szRet, szError)
t1 = Timer
WScript.Echo "Return      =" & ret
WScript.Echo "szAccess    =" & szAccess
WScript.Echo "szPassword  =" & szPassword
WScript.Echo "online      =" & online
WScript.Echo "szRet       =" & szRet
WScript.Echo "szError     =" & szError
WScript.Echo "elapsed(sec)=" & Format(t1 - t0, "0.00")"""
ws2["B33"] = vbs_code
ws2["B33"].font = Font(name="Consolas", size=10)
ws2["B33"].alignment = wrap_left
ws2.merge_cells("B33:F46")
ws2.row_dimensions[33].height = 15

ws2["B48"] = "実行: C:\\Windows\\System32\\cscript.exe //nologo test_aaverifyid.vbs (64bit 版 cscript を明示)"
ws2["B48"].font = normal_font
ws2.merge_cells("B48:F48")

ws2["B50"] = "6. テスト完了後の復元 (全テスト終了後に実施)"
ws2["B50"].font = section_font
ws2["B50"].fill = section_fill
ws2.merge_cells("B50:F50")

ws2["B51"] = "レジストリ : reg import sys148_backup.reg"
ws2["B52"] = "services  : copy services_backup.txt C:\\Windows\\System32\\drivers\\etc\\services"
for r in [51, 52]:
    ws2.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)


# =============================================================================
# Sheet 3: スタブサーバ
# =============================================================================
ws3 = wb.create_sheet("スタブサーバ")
ws3.sheet_view.showGridLines = False

ws3["B2"] = "スタブサーバ (stub_server.ps1)"
ws3["B2"].font = title_font
ws3.merge_cells("B2:F2")

ws3["B4"] = "1. 概要"
ws3["B4"].font = section_font
ws3["B4"].fill = section_fill
ws3.merge_cells("B4:F4")

ws3["B5"] = (
    "PowerShell 製の簡易スタブサーバ。AAVerifyID の異常系テスト (szRet=2/3/4) 用に、"
    "認証サーバの代替として同一 PC 上で動作する。起動時 -Mode パラメータで応答パターンを切替える。"
)
ws3["B5"].font = normal_font
ws3["B5"].alignment = wrap_left
ws3.merge_cells("B5:F7")
ws3.row_dimensions[5].height = 35

ws3["B9"] = "2. モード一覧"
ws3["B9"].font = section_font
ws3["B9"].fill = section_fill
ws3.merge_cells("B9:F9")

set_header_row(ws3, 10, ["Mode", "応答電文 (※LF = 0x0A)", "期待される szRet", "用途"], [22, 40, 18, 30])
modes = [
    ("normal-ack", 'ACK testuser, AAAAA, 1 + LF', '"0" (開局)', "正常応答 (開局)"),
    ("normal-ack-closed", 'ACK testuser, AAAAA, 0 + LF', '"0" (閉局)', "正常応答 (閉局)"),
    ("ng", "NG + LF", '"3"', "認証失敗応答"),
    ("malformed-commas", "ACK onlyone + LF", '"3"', "応答解析失敗 (カンマ不足)"),
    ("malformed-header", "XYZ foo,bar,1 + LF", '"3"', "応答解析失敗 (ACK 以外)"),
    ("silent", "(応答なし / accept のみ)", '"4"', "受信タイムアウト"),
]
for i, row in enumerate(modes):
    write_row(ws3, 11 + i, list(row),
              [wrap_center, wrap_left, wrap_center, wrap_left])

ws3["B18"] = "3. 起動手順"
ws3["B18"].font = section_font
ws3["B18"].fill = section_fill
ws3.merge_cells("B18:F18")

startup_steps = [
    "1. PowerShell を「管理者として実行」で開く",
    "2. stub_server.ps1 の配置フォルダに cd する",
    "3. .\\stub_server.ps1 -Mode <モード名>  を実行",
    "4. \"Listening on :50000...\" が表示されたら準備完了",
    "5. VBScript 側のテストを実行",
    "6. テスト完了後は Ctrl+C で停止",
]
for i, s in enumerate(startup_steps):
    ws3.cell(row=19 + i, column=2, value=s).font = normal_font
    ws3.merge_cells(start_row=19 + i, start_column=2, end_row=19 + i, end_column=6)

ws3["B26"] = "4. 起動例"
ws3["B26"].font = section_font
ws3["B26"].fill = section_fill
ws3.merge_cells("B26:F26")

examples = [
    "PS> .\\stub_server.ps1 -Mode normal-ack",
    "PS> .\\stub_server.ps1 -Mode ng",
    "PS> .\\stub_server.ps1 -Mode silent",
    "PS> .\\stub_server.ps1 -Mode normal-ack -Port 50000",
]
for i, s in enumerate(examples):
    ws3.cell(row=27 + i, column=2, value=s).font = Font(name="Consolas", size=10)
    ws3.merge_cells(start_row=27 + i, start_column=2, end_row=27 + i, end_column=6)

ws3["B32"] = "5. ファイアウォール"
ws3["B32"].font = section_font
ws3["B32"].fill = section_fill
ws3.merge_cells("B32:F32")

ws3["B33"] = (
    "初回起動時に Windows Defender ファイアウォールのダイアログが表示された場合は"
    " [アクセスを許可する] を選択する。"
)
ws3["B33"].font = normal_font
ws3["B33"].alignment = wrap_left
ws3.merge_cells("B33:F34")


# =============================================================================
# Sheet 4: テストケース一覧
# =============================================================================
ws4 = wb.create_sheet("テストケース一覧")

ws4["A1"] = "AAVerifyID テストケース一覧"
ws4["A1"].font = title_font
ws4.merge_cells("A1:I1")
ws4.row_dimensions[1].height = 25

headers = ["No", "分類", "テスト項目", "前提条件 (スタブ Mode 他)", "入力 user / pass",
           "期待 戻り値", "期待 szRet", "期待 szAccess / szPassword / online / szError",
           "実施手順概要"]
widths  = [  8,   14,    28,           36,                       20,
             10,         10,        42,
             32]
set_header_row(ws4, 2, headers, widths)
ws4.row_dimensions[2].height = 40

cases = [
    # (No, 分類, 項目, 前提, 入力, 戻値, szRet, 他出力, 手順)
    ("A-01", "正常系", "開局時の正常認証",
     "スタブ Mode=normal-ack、Host01=127.0.0.1",
     "DUMMY / DUMMY", "0", '"0"',
     'szAccess="AAAAA", szPassword=サーバ応答値, online=1, szError=""',
     "スタブ起動 → VBScript 実行 → 出力確認"),

    ("A-02", "正常系", "閉局時の正常認証",
     "スタブ Mode=normal-ack-closed",
     "DUMMY / DUMMY", "0", '"0"',
     'szAccess="AAAAA", online=0',
     "スタブ起動 → VBScript 実行 → online=0 確認"),

    ("A-03", "正常系", "フェイルオーバ成功",
     "Host01=192.0.2.1 (到達不可)、Host02=127.0.0.1、スタブ Mode=normal-ack",
     "DUMMY / DUMMY", "0", '"0"',
     'szAccess="AAAAA"',
     "レジストリ変更 → スタブ起動 → VBScript 実行"),

    ("A-04", "異常系", "初期化失敗",
     "HKLM\\SOFTWARE\\sys148 配下を一時削除 (事前 reg export /reg:64)",
     "DUMMY / DUMMY", "0", '"1"',
     'その他出力は空/0',
     "reg export → reg delete → VBScript 実行 → reg import で復元"),

    ("A-05", "異常系", "コネクション失敗",
     "スタブ未起動、Host01/02=127.0.0.1",
     "DUMMY / DUMMY", "0", '"2"',
     'szError に接続エラー情報',
     "スタブ停止状態で VBScript 実行"),

    ("A-06", "異常系", "NG 応答",
     "スタブ Mode=ng",
     "DUMMY / DUMMY", "0", '"3"',
     'szError に "NG" を含む',
     "スタブ起動 → VBScript 実行"),

    ("A-07", "異常系", "応答解析失敗 (カンマ不足)",
     "スタブ Mode=malformed-commas",
     "DUMMY / DUMMY", "0", '"3"',
     '—',
     "スタブ起動 → VBScript 実行"),

    ("A-08", "異常系", "応答解析失敗 (ヘッダ不正)",
     "スタブ Mode=malformed-header",
     "DUMMY / DUMMY", "0", '"3"',
     '—',
     "スタブ起動 → VBScript 実行"),

    ("A-09", "異常系", "受信タイムアウト (既定)",
     "スタブ Mode=silent、SocketTimeOut=5",
     "DUMMY / DUMMY", "0", '"4"',
     '所要時間 5±1 秒',
     "スタブ (silent) 起動 → VBScript 実行 → 経過時間確認"),

    ("A-10", "異常系", "受信タイムアウト (値変更)",
     "スタブ Mode=silent、SocketTimeOut=10",
     "DUMMY / DUMMY", "0", '"4"',
     '所要時間 10±1 秒',
     "レジストリ変更 → 新インスタンスで実行"),

    ("A-11", "境界", "user 空文字",
     "スタブ Mode=normal-ack",
     '"" / DUMMY', "0", '"0"',
     'スタブは常に ACK を返すため正常終了',
     "VBScript の user 引数を空文字に変更して実行"),

    ("A-12", "境界", "pass 空文字",
     "スタブ Mode=normal-ack",
     'DUMMY / ""', "0", '"0"',
     '同上',
     "VBScript の pass 引数を空文字に変更して実行"),

    ("A-13", "境界", "長大文字列 (1024 文字)",
     "スタブ Mode=normal-ack",
     'String(1024,"a") / DUMMY', "0", '"0"',
     'コマンド組立で例外が出ないこと',
     "長大文字列を VBScript で生成して実行"),

    ("A-14", "境界", "全角文字",
     "スタブ Mode=normal-ack",
     '"テスト" / DUMMY', "0", '"0"',
     'エンコード挙動の確認',
     "VBScript で全角引数を渡して実行"),

    ("A-15", "不変性", "関数戻り値 = 0",
     "A-01 〜 A-14 全ケース共通",
     "—", "0", '—',
     'すべてのケースで関数戻り値が 0 であること',
     "各ケース実施時に Return=0 を確認"),
]

for i, case in enumerate(cases):
    r = 3 + i
    aligns = [wrap_center, wrap_center, wrap_left, wrap_left, wrap_center,
              wrap_center, wrap_center, wrap_left, wrap_left]
    write_row(ws4, r, list(case), aligns)
    ws4.row_dimensions[r].height = 48

ws4.freeze_panes = "A3"


# =============================================================================
# Sheet 5: 実施記録
# =============================================================================
ws5 = wb.create_sheet("実施記録")

ws5["A1"] = "AAVerifyID テスト 実施記録"
ws5["A1"].font = title_font
ws5.merge_cells("A1:G1")
ws5.row_dimensions[1].height = 25

rec_headers = ["No", "テスト項目", "実施日", "実施者", "結果 (OK/NG)", "実測値 (szRet 等)", "備考"]
rec_widths  = [8, 28, 14, 12, 12, 32, 32]
set_header_row(ws5, 2, rec_headers, rec_widths)
ws5.row_dimensions[2].height = 30

for i, case in enumerate(cases):
    r = 3 + i
    values = [case[0], case[2], "", "", "", "", ""]
    aligns = [wrap_center, wrap_left, wrap_center, wrap_center, wrap_center, wrap_left, wrap_left]
    write_row(ws5, r, values, aligns)
    ws5.row_dimensions[r].height = 24

ws5.freeze_panes = "A3"

ws5.cell(row=len(cases) + 5, column=1, value="備考欄")
ws5.cell(row=len(cases) + 5, column=1).font = subtitle_font
ws5.cell(row=len(cases) + 6, column=1,
         value="・環境変更を伴うケース (A-04, A-10) は実施後にレジストリを復元すること")
ws5.cell(row=len(cases) + 7, column=1,
         value="・スタブサーバの起動・停止を誤ると他ケースに影響するため、1 ケース終了ごとに Ctrl+C で停止推奨")
for r in [len(cases) + 6, len(cases) + 7]:
    ws5.cell(row=r, column=1).font = normal_font
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)


# =============================================================================
out = "/home/ubuntu/work/BOJClientVB_UnitTestSpec_AAVerifyID_v1.5.xlsx"
wb.save(out)
print(f"OK: {out}")
